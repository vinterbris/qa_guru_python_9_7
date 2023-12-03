import os.path, shutil, pytest, csv
from io import BytesIO
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook


@pytest.fixture
def folders():
    if not os.path.exists('resources'):
        os.mkdir('resources')
    if not os.path.exists('tmp'):
        os.mkdir('tmp')

@pytest.fixture
def archiving():
    if not os.path.exists('resources/test_archive.zip'):
        shutil.make_archive('test_archive', 'zip', 'tmp')
        shutil.move('test_archive.zip', 'resources/test_archive.zip')


def test_csv(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('ubuntu.csv')
        assert archive.getinfo('ubuntu.csv').file_size == 2583


def test_pdf(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('Storytelling.pdf')
        reader = PdfReader(BytesIO(file))
        assert archive.getinfo('Storytelling.pdf').file_size == 40365
        assert len(reader.pages) == 1
        assert "I tried to kiss a girl and head butt her instead" in reader.pages[0].extract_text()


def test_xlsx(folders, archiving):
    with ZipFile('resources/test_archive.zip') as archive:
        file = archive.read('clinics.xlsx')
        workbook = load_workbook(BytesIO(file))
        sheet = workbook.active
        assert archive.getinfo('clinics.xlsx').file_size == 11801
        assert len(workbook.sheetnames) == 1
        assert sheet.max_row == 63
        assert sheet.max_column == 4
        assert sheet.cell(row=62, column=2).value == 'Клиника Эксперт'